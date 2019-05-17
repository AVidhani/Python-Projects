# from openpyxl import load_workbook
import openpyxl
import datetime, string, sys



#Import windows api for MS outlook
import win32com.client as win32

textfile = "messageContent.txt"
msgContents=""
in_file = open(textfile,"r")

for line in in_file:
    msgContents+=line
in_file.close()

workbook = openpyxl.load_workbook('sample.xlsx')
first_sheet = workbook.get_sheet_names()[1]
worksheet = workbook.get_sheet_by_name(first_sheet)
email = []
name = []

for row in worksheet.iter_rows(): #iterate through rows in the worksheet
    #print(row)
    num = 0
    # check out the last row
    for cell in row:#iterate through cell in a row
        if num == 3:
            name += [cell.value]
        if num == 4:
            # print(cell.value)
            if cell.value not in email: # do not add this email if it is already there in the email array
                email += [cell.value]

        if num == 5:
            parts= str(cell.value).split()
            print("Spreadsheet date is " + str(parts[0]) + " and current date time value is " + str(datetime.date.today()))
            if parts[0]!="Open":
                xlDate = datetime.datetime.strptime(str(parts[0]), '%Y-%m-%d')
                d = datetime.date.today()
                if d.day != xlDate.day or d.month != xlDate.month:
                    email.pop()
                    name.pop()
            else:
                email.pop()
                name.pop()
        num += 1

print("length of email list = " + str(len(email)) + " " + email[0] + " and length of names list = " + str(len(name)) + " " + name[0])
workbook.close()
for i in range(len(name)):
    print("Hi, I am in loop!")
    salutation="Dear " + name[i] + ",\n\n"
    message=salutation+msgContents
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = email[i]
    mail.Subject = 'Sending email from python script'
    mail.Body = message
    mail.Send()
